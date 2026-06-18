"""
Riport generátor – Excel és opcionális Parquet export.

Az Excel 5 lapot tartalmaz:
    ✅ Releváns         – score ≥ threshold + feedback jóváhagyott
    ⚠️  Felülvizsgálandó – Döntés dropdown kitöltendő
    ❌ Zaj              – spot-check célokra
    📄 Összes találat   – teljes adatsor
    📊 Összefoglaló     – kulcsszó statisztikák

A Parquet export (DS csapatnak) opcionális: csak ha a pyarrow telepítve van.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ikon.config import ExportConfig
from ikon.keywords import ALL_KEYWORDS, KEYWORD_TIER
from ikon.models import FeedbackDecision, RawArticle, ScoredArticle

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Oszlop konfiguráció
# ---------------------------------------------------------------------------

_COLS_REVIEW = [
    "Döntés", "datum", "ido", "cim", "forras", "forras_tipus",
    "lead", "url", "matching_keywords", "legjobb_tier", "score", "miért", "Megjegyzés",
]

_COLS_NORMAL = [
    "datum", "ido", "cim", "forras", "forras_tipus",
    "lead", "url", "matching_keywords", "legjobb_tier", "score", "miért",
    "kategoria", "feedback_döntés",
]

_WIDTHS: dict[str, int] = {
    "Döntés": 18, "datum": 12, "ido": 7, "cim": 55, "forras": 16,
    "forras_tipus": 14, "lead": 65, "url": 45, "matching_keywords": 32,
    "legjobb_tier": 16, "score": 7, "miért": 55, "kategoria": 16,
    "Megjegyzés": 30, "feedback_döntés": 20,
}

_HEADER_COLORS: dict[str, str] = {
    "Releváns": "1F4E79",
    "Felülvizsgálandó": "7F5F00",
    "Zaj": "5C2B00",
    "Összes találat": "2E4057",
    "Összefoglaló": "1B4332",
}


# ---------------------------------------------------------------------------
# Modell → DataFrame konverzió
# ---------------------------------------------------------------------------

def _feedback_label(article: ScoredArticle) -> str:
    if article.feedback_decision == FeedbackDecision.RELEVANT:
        return "✅ Jóváhagyva"
    if article.feedback_decision == FeedbackDecision.NOT_RELEVANT:
        return "❌ Elutasítva"
    return ""


def scored_to_df(articles: list[ScoredArticle]) -> pd.DataFrame:
    """Konvertálja a ScoredArticle listát a riporthoz szükséges DataFrame-mé."""
    if not articles:
        return pd.DataFrame(columns=_COLS_NORMAL)

    rows = [
        {
            "datum": a.published_date,
            "ido": a.published_time,
            "cim": a.title,
            "forras": a.source,
            "forras_tipus": a.source_type.value,
            "lead": a.excerpt,
            "url": a.url,
            "matching_keywords": ", ".join(sorted(a.matched_keywords)),
            "legjobb_tier": a.best_tier.value,
            "score": a.score,
            "miért": a.score_reason,
            "kategoria": a.effective_category.value,
            "feedback_döntés": _feedback_label(a),
        }
        for a in articles
    ]
    df = pd.DataFrame(rows)
    # Rendezés dátum + pontszám szerint
    df["_datum_sort"] = pd.to_datetime(df["datum"], format="%Y.%m.%d", errors="coerce")
    df = df.sort_values(["_datum_sort", "score"], ascending=[False, False])
    return df.drop(columns=["_datum_sort"])


def raw_to_df(articles: list[RawArticle]) -> pd.DataFrame:
    if not articles:
        return pd.DataFrame()
    return pd.DataFrame([
        {
            "keyword": a.matched_keyword,
            "tier": a.keyword_tier,
            "url": a.url,
        }
        for a in articles
    ])


# ---------------------------------------------------------------------------
# Összefoglaló lap
# ---------------------------------------------------------------------------

def build_summary(
    raw_df: pd.DataFrame,
    scored_df: pd.DataFrame,
) -> pd.DataFrame:
    """Kulcsszó-szintű statisztikák. Minden kulcsszó megjelenik (0-találatosak is)."""
    rows = []
    for kw in ALL_KEYWORDS:
        kw_total = int((raw_df["keyword"] == kw).sum()) if not raw_df.empty else 0
        kw_relevant = 0
        if not scored_df.empty and "matching_keywords" in scored_df.columns:
            mask = (
                scored_df["matching_keywords"].str.contains(re.escape(kw), na=False)
                & (scored_df["kategoria"] == "releváns")
            )
            kw_relevant = int(mask.sum())
        rows.append({
            "kulcsszó": kw,
            "tier": KEYWORD_TIER.get(kw, "?"),
            "összes_találat": kw_total,
            "releváns_találat": kw_relevant,
            "zaj_%": round(100 * (1 - kw_relevant / kw_total), 1) if kw_total else "-",
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Sheet szétválasztás
# ---------------------------------------------------------------------------

def _split_sheets(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    return {
        "Releváns": df[df["kategoria"] == "releváns"].copy(),
        "Felülvizsgálandó": df[df["kategoria"] == "felülvizsgálandó"].copy(),
        "Zaj": df[df["kategoria"] == "zaj"].copy(),
        "Összes találat": df.copy(),
    }


# ---------------------------------------------------------------------------
# Excel formázás
# ---------------------------------------------------------------------------

def _style_sheet(ws, df: pd.DataFrame, sheet_name: str) -> None:
    color = _HEADER_COLORS.get(sheet_name, "1F4E79")
    header_fill = PatternFill("solid", fgColor=color)
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = _WIDTHS.get(col_name, 14)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _color_score_col(ws, df: pd.DataFrame) -> None:
    if "score" not in df.columns:
        return
    sc_col = list(df.columns).index("score") + 1
    for r in range(2, len(df) + 2):
        val = ws.cell(row=r, column=sc_col).value or 0
        if val >= 40:
            fill = PatternFill("solid", fgColor="C6EFCE")
        elif val >= 20:
            fill = PatternFill("solid", fgColor="FFEB9C")
        elif val >= 8:
            fill = PatternFill("solid", fgColor="FCE4D6")
        else:
            fill = PatternFill("solid", fgColor="FFC7CE")
        ws.cell(row=r, column=sc_col).fill = fill


def _add_hyperlinks(ws, df: pd.DataFrame) -> None:
    if "url" not in df.columns or "cim" not in df.columns:
        return
    url_col = list(df.columns).index("url") + 1
    cim_col = list(df.columns).index("cim") + 1
    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        url = str(row.get("url", ""))
        if url.startswith("http"):
            cell = ws.cell(row=r_idx, column=cim_col)
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")


def _add_döntés_dropdown(ws, nrows: int) -> None:
    dv = DataValidation(
        type="list",
        formula1='"✅ Releváns,❌ Nem releváns"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Kérjük válasszon a listából"
    dv.errorTitle = "Érvénytelen érték"
    ws.add_data_validation(dv)
    if nrows > 0:
        dv.add(f"A2:A{nrows + 1}")
    for r in range(2, nrows + 2):
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="FFF2CC")


def _prepare_review_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["Döntés"] = out.get("feedback_döntés", "")
    out["Megjegyzés"] = ""
    cols = [c for c in _COLS_REVIEW if c in out.columns or c in ("Döntés", "Megjegyzés")]
    for c in cols:
        if c not in out.columns:
            out[c] = ""
    return out[cols]


def _select_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    available = [c for c in cols if c in df.columns]
    return df[available]


# ---------------------------------------------------------------------------
# Publikus export függvények
# ---------------------------------------------------------------------------

def export_excel(
    articles: list[ScoredArticle],
    raw_articles: list[RawArticle],
    output_path: Path | str,
    cfg: Optional[ExportConfig] = None,
) -> None:
    """Generálja az 5-lapos Excel riportot."""
    cfg = cfg or ExportConfig()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    scored_df = scored_to_df(articles)
    raw_df = raw_to_df(raw_articles)
    sheets = _split_sheets(scored_df)
    summary = build_summary(raw_df, scored_df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 1. Releváns
        rel_df = _select_cols(sheets["Releváns"], _COLS_NORMAL)
        rel_df.to_excel(writer, sheet_name="Releváns", index=False)
        ws = writer.sheets["Releváns"]
        _style_sheet(ws, rel_df, "Releváns")
        _color_score_col(ws, rel_df)
        _add_hyperlinks(ws, rel_df)

        # 2. Felülvizsgálandó (dropdown)
        rev_df = _prepare_review_df(sheets["Felülvizsgálandó"])
        rev_df.to_excel(writer, sheet_name="Felülvizsgálandó", index=False)
        ws = writer.sheets["Felülvizsgálandó"]
        _style_sheet(ws, rev_df, "Felülvizsgálandó")
        _color_score_col(ws, rev_df)
        _add_hyperlinks(ws, rev_df)
        _add_döntés_dropdown(ws, len(rev_df))

        # 3. Zaj
        zaj_df = _select_cols(sheets["Zaj"], _COLS_NORMAL)
        zaj_df.to_excel(writer, sheet_name="Zaj", index=False)
        ws = writer.sheets["Zaj"]
        _style_sheet(ws, zaj_df, "Zaj")
        _color_score_col(ws, zaj_df)

        # 4. Összes találat
        all_df = _select_cols(sheets["Összes találat"], _COLS_NORMAL)
        all_df.to_excel(writer, sheet_name="Összes találat", index=False)
        ws = writer.sheets["Összes találat"]
        _style_sheet(ws, all_df, "Összes találat")
        _color_score_col(ws, all_df)

        # 5. Összefoglaló
        summary.to_excel(writer, sheet_name="Összefoglaló", index=False)
        ws = writer.sheets["Összefoglaló"]
        _style_sheet(ws, summary, "Összefoglaló")

    rl = len(sheets["Releváns"])
    fv = len(sheets["Felülvizsgálandó"])
    zj = len(sheets["Zaj"])

    logger.info("Excel mentve: %s", output_path)
    logger.info("  Releváns: %d | Felülvizsgálandó: %d | Zaj: %d | Összes: %d",
                rl, fv, zj, len(articles))


def export_parquet(
    articles: list[ScoredArticle],
    output_dir: Path | str,
    run_id: str,
) -> Optional[Path]:
    """Parquet export a DS csapatnak. Csak ha pyarrow telepítve van."""
    try:
        import pyarrow  # noqa: F401 – csak ellenőrzés
    except ImportError:
        logger.debug("pyarrow nem elérhető – Parquet export kihagyva")
        return None

    df = scored_to_df(articles)
    df["run_id"] = run_id
    out = Path(output_dir) / f"articles_{run_id}.parquet"
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out, index=False)
    logger.info("Parquet mentve: %s (%d sor)", out, len(df))
    return out
