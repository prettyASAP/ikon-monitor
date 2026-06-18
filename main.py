"""
Hírkereső Média Monitoring Pipeline

Futtatás:
  python main.py                                          # friss scraping
  python main.py --from-cache output/raw_YYYY-MM-DD.csv  # cached adat újrapontozása
  python main.py --apply-feedback output/hirek_YYYY-MM-DD.xlsx  # emberi döntések mentése
"""
import argparse
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from scraper import scrape_all
from cleaner import deduplicate_and_score, split_sheets, build_summary
from reviewer import (
    load_feedback, save_feedback,
    read_decisions_from_excel, apply_feedback_to_scored,
    print_feedback_stats,
)

# ---------------------------------------------------------------------------
# Oszlop definíciók laponként
# ---------------------------------------------------------------------------

COLS_REVIEW = [
    "Döntés", "datum", "ido", "cim", "forras", "forras_tipus",
    "lead", "url", "matching_keywords", "legjobb_tier", "score", "miért", "Megjegyzés",
]

COLS_NORMAL = [
    "datum", "ido", "cim", "forras", "forras_tipus",
    "lead", "url", "matching_keywords", "legjobb_tier", "score", "miért",
    "kategoria", "feedback_döntés",
]

WIDTHS = {
    "Döntés": 18, "datum": 12, "ido": 7, "cim": 55, "forras": 16,
    "forras_tipus": 14, "lead": 65, "url": 45, "matching_keywords": 32,
    "legjobb_tier": 16, "score": 7, "miért": 55, "kategoria": 16,
    "Megjegyzés": 30, "feedback_döntés": 20,
}

HEADER_COLORS = {
    "Releváns": "1F4E79",
    "Felülvizsgálandó": "7F5F00",
    "Zaj": "5C2B00",
    "Összes találat": "2E4057",
    "Összefoglaló": "1B4332",
}

# ---------------------------------------------------------------------------
# Excel formázás
# ---------------------------------------------------------------------------

def _style_sheet(ws, df: pd.DataFrame, sheet_name: str):
    color = HEADER_COLORS.get(sheet_name, "1F4E79")
    header_fill = PatternFill("solid", fgColor=color)
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = WIDTHS.get(col_name, 14)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _color_score_col(ws, df: pd.DataFrame):
    if "score" not in df.columns:
        return
    sc_col = list(df.columns).index("score") + 1
    for r in range(2, len(df) + 2):
        val = ws.cell(row=r, column=sc_col).value or 0
        if val >= 40:
            fill = PatternFill("solid", fgColor="C6EFCE")
        elif val >= 20:
            fill = PatternFill("solid", fgColor="FFEB9C")
        elif val >= 8:
            fill = PatternFill("solid", fgColor="FCE4D6")
        else:
            fill = PatternFill("solid", fgColor="FFC7CE")
        ws.cell(row=r, column=sc_col).fill = fill


def _add_hyperlinks(ws, df: pd.DataFrame):
    if "url" not in df.columns or "cim" not in df.columns:
        return
    url_col = list(df.columns).index("url") + 1
    cim_col = list(df.columns).index("cim") + 1
    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        url = str(row.get("url", ""))
        if url.startswith("http"):
            cim_cell = ws.cell(row=r_idx, column=cim_col)
            cim_cell.hyperlink = url
            cim_cell.font = Font(color="0563C1", underline="single")


def _add_döntés_dropdown(ws, nrows: int):
    """Data Validation dropdown a Döntés oszlopra (A oszlop)."""
    dv = DataValidation(
        type="list",
        formula1='"✅ Releváns,❌ Nem releváns"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Kérjük válasszon a listából"
    dv.errorTitle = "Érvénytelen érték"
    ws.add_data_validation(dv)
    if nrows > 0:
        dv.add(f"A2:A{nrows + 1}")
    # Döntés oszlop sárga háttér (kitöltésre vár)
    for r in range(2, nrows + 2):
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="FFF2CC")


def _prepare_review_df(df: pd.DataFrame) -> pd.DataFrame:
    """Felülvizsgálandó laphoz: Döntés és Megjegyzés oszlop hozzáadása."""
    out = df.copy()
    if "feedback_döntés" in out.columns:
        out["Döntés"] = out["feedback_döntés"]
    else:
        out["Döntés"] = ""
    out["Megjegyzés"] = ""
    cols = [c for c in COLS_REVIEW if c in out.columns or c in ("Döntés", "Megjegyzés")]
    for c in cols:
        if c not in out.columns:
            out[c] = ""
    return out[cols]


# ---------------------------------------------------------------------------
# Excel írás
# ---------------------------------------------------------------------------

def write_excel(scored_df: pd.DataFrame, raw_df: pd.DataFrame, output_path: str):
    sheets = split_sheets(scored_df)
    summary = build_summary(raw_df, scored_df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # 1. Releváns
        rel_df = sheets["Releváns"][[c for c in COLS_NORMAL if c in sheets["Releváns"].columns]]
        rel_df.to_excel(writer, sheet_name="Releváns", index=False)
        ws = writer.sheets["Releváns"]
        _style_sheet(ws, rel_df, "Releváns")
        _color_score_col(ws, rel_df)
        _add_hyperlinks(ws, rel_df)

        # 2. Felülvizsgálandó (dropdown + hyperlink)
        rev_src = sheets["Felülvizsgálandó"]
        rev_df = _prepare_review_df(rev_src)
        rev_df.to_excel(writer, sheet_name="Felülvizsgálandó", index=False)
        ws = writer.sheets["Felülvizsgálandó"]
        _style_sheet(ws, rev_df, "Felülvizsgálandó")
        _color_score_col(ws, rev_df)
        _add_hyperlinks(ws, rev_df)
        _add_döntés_dropdown(ws, len(rev_df))

        # 3. Zaj (spot-check)
        zaj_df = sheets["Zaj"][[c for c in COLS_NORMAL if c in sheets["Zaj"].columns]]
        zaj_df.to_excel(writer, sheet_name="Zaj", index=False)
        ws = writer.sheets["Zaj"]
        _style_sheet(ws, zaj_df, "Zaj")
        _color_score_col(ws, zaj_df)

        # 4. Összes találat
        all_df = sheets["Összes találat"][[c for c in COLS_NORMAL if c in sheets["Összes találat"].columns]]
        all_df.to_excel(writer, sheet_name="Összes találat", index=False)
        ws = writer.sheets["Összes találat"]
        _style_sheet(ws, all_df, "Összes találat")
        _color_score_col(ws, all_df)

        # 5. Összefoglaló
        summary.to_excel(writer, sheet_name="Összefoglaló", index=False)
        ws = writer.sheets["Összefoglaló"]
        _style_sheet(ws, summary, "Összefoglaló")

    rl, fv, zj = len(sheets["Releváns"]), len(sheets["Felülvizsgálandó"]), len(sheets["Zaj"])
    print(f"\nExcel mentve: {output_path}")
    print(f"  ✅ Releváns:          {rl} cikk")
    print(f"  ⚠️  Felülvizsgálandó:  {fv} cikk  ← Döntés dropdown kitöltendő")
    print(f"  ❌ Zaj:               {zj} cikk")
    print(f"  📄 Összes (unique):   {len(scored_df)} cikk")
    print(f"  📥 Nyers találatok:   {len(raw_df)} sor (duplikátumokkal)")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Hírkereső média monitoring v2")
    parser.add_argument("--from-cache", metavar="CSV", help="Meglévő nyers CSV betöltése")
    parser.add_argument(
        "--apply-feedback", metavar="XLSX",
        help="Emberi döntések beolvasása a Felülvizsgálandó lapból → feedback.json",
    )
    args = parser.parse_args()

    Path("output").mkdir(exist_ok=True)
    today = date.today().isoformat()

    # --- Csak feedback alkalmazása ---
    if args.apply_feedback:
        print(f"Feedback beolvasása: {args.apply_feedback}")
        decisions = read_decisions_from_excel(args.apply_feedback)
        if decisions:
            save_feedback(decisions)
            print_feedback_stats(decisions)
        else:
            print("Nem találtam kitöltött döntést a Felülvizsgálandó lapon.")
        return

    # --- Scraping vagy cache ---
    if args.from_cache:
        print(f"Cache betöltése: {args.from_cache}")
        raw_df = pd.read_csv(args.from_cache, encoding="utf-8-sig")
    else:
        print("Scraping indul...\n")
        raw_df = pd.DataFrame(scrape_all())
        raw_path = f"output/raw_{today}.csv"
        raw_df.to_csv(raw_path, index=False, encoding="utf-8-sig")
        print(f"\nNyers adatok mentve: {raw_path}")

    print(f"\nNyers találatok: {len(raw_df)} sor")
    print("Deduplikálás és pontozás...")
    scored_df = deduplicate_and_score(raw_df)

    # Feedback alkalmazása (ha van)
    feedback = load_feedback()
    if feedback:
        print(f"Feedback betöltve: {len(feedback)} URL a tárházból")
        scored_df = apply_feedback_to_scored(scored_df, feedback)

    output_path = f"output/hirek_{today}.xlsx"
    write_excel(scored_df, raw_df, output_path)


if __name__ == "__main__":
    main()
